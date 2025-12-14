"""
API Views for CPSU Virtual Health Assistant
Implements all endpoints for student and clinic staff
"""

from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate, get_user_model
from django.utils import timezone
from django.db.models import Count, Q
from datetime import timedelta
import uuid
import logging

from .models import SymptomRecord, HealthInsight, ChatSession, ConsentLog, AuditLog, DepartmentStats, EmergencyAlert, Medication, MedicationLog, FollowUp
from .serializers import (
    UserRegistrationSerializer, UserProfileSerializer,
    SymptomRecordSerializer, SymptomSubmissionSerializer,
    DiseasePredictionSerializer, HealthInsightSerializer,
    ChatSessionSerializer, ChatMessageSerializer,
    ConsentLogSerializer, AuditLogSerializer,
    DepartmentStatsSerializer, DashboardStatsSerializer,
    EmergencyAlertSerializer, EmergencyTriggerSerializer,
    MedicationSerializer, MedicationCreateSerializer, MedicationLogSerializer,
    FollowUpSerializer, FollowUpResponseSerializer
)
from .permissions import IsStudent, IsClinicStaff, IsOwnerOrStaff, CanModifyProfile, HasDataConsent
from .ml_service import get_ml_predictor

logger = logging.getLogger(__name__)
from .llm_service import AIInsightGenerator
from .rasa_service import RasaChatService

# Get singleton instances
ml_predictor = get_ml_predictor()
ai_generator = AIInsightGenerator()
rasa_service = RasaChatService()

User = get_user_model()


# ============================================================================
# Authentication Views
# ============================================================================

@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """
    Register new user (student or staff)
    POST /api/auth/register/
    """
    serializer = UserRegistrationSerializer(data=request.data)
    
    if serializer.is_valid():
        # Determine role based on registration context
        # In production, staff registration would be admin-only
        role = request.data.get('role', 'student')
        if role not in ['student', 'staff']:
            role = 'student'
        
        user = serializer.save(role=role)
        
        # Create auth token
        token, created = Token.objects.get_or_create(user=user)
        
        # Log consent if given
        if user.data_consent_given:
            ConsentLog.objects.create(
                user=user,
                action='granted',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )
        
        return Response({
            'token': token.key,
            'user': UserProfileSerializer(user).data,
            'message': 'Registration successful'
        }, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    """
    User login with school_id and password
    POST /api/auth/login/
    """
    school_id = request.data.get('school_id')
    password = request.data.get('password')
    
    if not school_id or not password:
        return Response(
            {'error': 'School ID and password are required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(username=school_id, password=password)
    
    if user:
        token, created = Token.objects.get_or_create(user=user)
        
        return Response({
            'token': token.key,
            'user': UserProfileSerializer(user).data,
            'message': 'Login successful'
        })
    
    return Response(
        {'error': 'Invalid credentials'},
        status=status.HTTP_401_UNAUTHORIZED
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_user(request):
    """
    User logout (delete auth token)
    POST /api/auth/logout/
    """
    request.user.auth_token.delete()
    return Response({'message': 'Logout successful'})


# ============================================================================
# User Profile Views
# ============================================================================

class UserProfileView(generics.RetrieveUpdateAPIView):
    """
    Get or update user profile
    GET/PUT/PATCH /api/profile/
    """
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated, CanModifyProfile]
    
    def get_object(self):
        return self.request.user
    
    def update(self, request, *args, **kwargs):
        # Check for immutable fields
        immutable_fields = {'school_id', 'role'}
        if any(field in request.data for field in immutable_fields):
            return Response(
                {'error': 'Cannot modify school_id or role'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().update(request, *args, **kwargs)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_consent(request):
    """
    Update data consent status
    POST /api/profile/consent/
    """
    consent_given = request.data.get('data_consent_given')
    
    if consent_given is None:
        return Response(
            {'error': 'data_consent_given field is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = request.user
    previous_consent = user.data_consent_given
    user.data_consent_given = consent_given
    
    if consent_given and not user.consent_date:
        user.consent_date = timezone.now()
    
    user.save()
    
    # Log consent change
    action = 'granted' if consent_given else 'revoked'
    if previous_consent == consent_given:
        action = 'updated'
    
    ConsentLog.objects.create(
        user=user,
        action=action,
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
    )
    
    return Response({
        'message': f'Consent {action}',
        'data_consent_given': user.data_consent_given,
        'consent_date': user.consent_date
    })


# ============================================================================
# Symptom & ML Views
# ============================================================================

class SymptomRecordViewSet(viewsets.ModelViewSet):
    """
    Symptom record CRUD operations
    Students can only access their own records
    Staff can access all records
    """
    serializer_class = SymptomRecordSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrStaff]
    
    def get_queryset(self):
        user = self.request.user
        
        if user.role == 'staff':
            # Staff can see all records
            queryset = SymptomRecord.objects.all().select_related('student')
        else:
            # Students see only their own
            queryset = SymptomRecord.objects.filter(student=user)
        
        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        
        return queryset.order_by('-created_at')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def submit_symptoms(request):
    """
    Submit symptoms and get disease prediction
    POST /api/symptoms/submit/
    """
    serializer = SymptomSubmissionSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        # Get ML prediction
        predictor = get_ml_predictor()
        prediction_result = predictor.predict(data['symptoms'])
        
        # Create symptom record
        record = SymptomRecord.objects.create(
            student=request.user,
            symptoms=data['symptoms'],
            duration_days=data['duration_days'],
            severity=data['severity'],
            on_medication=data.get('on_medication', False),
            medication_adherence=data.get('medication_adherence'),
            predicted_disease=prediction_result['predicted_disease'],
            confidence_score=prediction_result['confidence_score'],
            top_predictions=prediction_result['top_predictions'],
            is_communicable=prediction_result['is_communicable'],
            is_acute=prediction_result['is_acute'],
            icd10_code=prediction_result['icd10_code']
        )
        
        # Check referral criteria
        record.check_referral_criteria()
        record.save()
        
        # Auto-create follow-up (3 days from now)
        followup = FollowUp.create_from_symptom(record, days_ahead=3)
        
        # Prepare response
        response_data = {
            'record_id': str(record.id),
            'prediction': prediction_result,
            'requires_referral': record.requires_referral,
            'referral_message': 'You have reported symptoms 5+ times in the past 30 days. Please visit the clinic for evaluation.' if record.requires_referral else None,
            'followup_scheduled': followup.scheduled_date.isoformat()
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        return Response(
            {'error': f'Prediction failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_available_symptoms(request):
    """
    Get list of all symptoms the ML model recognizes
    GET /api/symptoms/available/
    """
    try:
        predictor = get_ml_predictor()
        symptoms = predictor.get_available_symptoms()
        
        return Response({
            'count': len(symptoms),
            'symptoms': sorted(symptoms)
        })
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================================
# AI Chat & Insights Views
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def start_chat_session(request):
    """
    Start a new AI chat session
    POST /api/chat/start/
    """
    language = request.data.get('language', 'english')
    
    session = ChatSession.objects.create(
        student=request.user,
        language=language
    )
    
    return Response({
        'session_id': str(session.id),
        'message': 'Chat session started',
        'language': language
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def send_chat_message(request):
    """
    Send message in AI chat (real-time, not stored)
    POST /api/chat/message/
    """
    serializer = ChatMessageSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    message = data['message']
    language = data.get('language', 'english')
    session_id = data.get('session_id')
    
    try:
        # Verify session exists and belongs to user
        if session_id:
            session = ChatSession.objects.get(id=session_id, student=request.user)
        else:
            return Response(
                {'error': 'session_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # HYBRID FLOW: Rasa handles conversation → Django provides ML predictions
        # LLM only used as fallback when Rasa fails
        
        # Step 1: Send message to Rasa
        rasa_response = rasa_service.send_message(
            message=message,
            sender_id=str(session_id),
            metadata={
                'language': language,
                'user_id': str(request.user.id),
                'django_api': request.build_absolute_uri('/api/')  # Rasa can call back
            }
        )
        
        # Step 2: Check if we should use LLM fallback
        if rasa_service.should_use_llm_fallback(rasa_response):
            # LLM Fallback: Only when Rasa completely fails
            logger.warning(f"Using LLM fallback (Rasa {'unavailable' if not rasa_service.is_available() else 'low confidence'})")
            try:
                response_text = ai_generator.generate_chat_response(
                    message=message,
                    context={'language': language, 'session_id': str(session_id), 'rasa_failed': True}
                )
                # Ensure response is not None or empty
                if not response_text or not response_text.strip():
                    raise ValueError("LLM returned empty response")
            except Exception as llm_error:
                logger.error(f"LLM fallback failed: {llm_error}")
                # Ultimate fallback - hardcoded response
                response_text = "Thank you for your message. I'm experiencing technical difficulties. Please consult with our clinic staff for proper evaluation of your symptoms."
            
            response_source = "llm_fallback"
            buttons = []
            
            # Try to extract symptoms from message and get ML prediction for LLM fallback
            diagnosis_data = None
            try:
                predictor = get_ml_predictor()
                available_symptoms = predictor.get_available_symptoms()
                
                # Simple symptom extraction from message
                message_lower = message.lower().replace(' ', '_')
                extracted_symptoms = [s for s in available_symptoms if s in message_lower]
                
                # If we found symptoms, get ML prediction
                if extracted_symptoms:
                    prediction = predictor.predict(extracted_symptoms)
                    diagnosis_data = {
                        'symptoms': extracted_symptoms,
                        'predicted_disease': prediction.get('predicted_disease'),
                        'confidence': prediction.get('confidence_score', 0.0),
                        'top_predictions': prediction.get('top_predictions', []),
                        'is_communicable': prediction.get('is_communicable', False),
                        'is_acute': prediction.get('is_acute', False),
                        'icd10_code': prediction.get('icd10_code', ''),
                        'severity': 'moderate',
                        'duration_days': 1
                    }
                    logger.info(f"LLM fallback: extracted {len(extracted_symptoms)} symptoms, predicted {prediction.get('predicted_disease')}")
            except Exception as extract_error:
                logger.warning(f"Could not extract symptoms from LLM fallback message: {extract_error}")
        else:
            # Use Rasa response (Rasa handles conversation flow)
            response_text = rasa_response['text']
            response_source = "rasa"
            buttons = rasa_response.get('buttons', [])
            
            # Check if Rasa provided diagnosis data
            diagnosis_data = rasa_response.get('custom', {}).get('diagnosis')
        
        # Step 3: Save symptom record if diagnosis was provided
        record_id = None
        if diagnosis_data and diagnosis_data.get('predicted_disease'):
            try:
                # Extract symptoms from diagnosis data or session metadata
                symptoms = diagnosis_data.get('symptoms', [])
                
                # Convert severity string to integer (1=Mild, 2=Moderate, 3=Severe)
                severity_map = {'mild': 1, 'moderate': 2, 'severe': 3}
                severity_value = diagnosis_data.get('severity', 'moderate')
                if isinstance(severity_value, str):
                    severity_int = severity_map.get(severity_value.lower(), 2)  # Default to moderate (2)
                else:
                    severity_int = int(severity_value) if severity_value else 2
                
                # Create symptom record for history tracking
                record = SymptomRecord.objects.create(
                    student=request.user,
                    symptoms=symptoms,
                    duration_days=diagnosis_data.get('duration_days', 1),
                    severity=severity_int,
                    predicted_disease=diagnosis_data['predicted_disease'],
                    confidence_score=diagnosis_data.get('confidence', 0.0),
                    top_predictions=diagnosis_data.get('top_predictions', []),
                    is_communicable=diagnosis_data.get('is_communicable', False),
                    is_acute=diagnosis_data.get('is_acute', False),
                    icd10_code=diagnosis_data.get('icd10_code', '')
                )
                
                # Check referral criteria
                record.check_referral_criteria()
                record.save()
                
                # Auto-create follow-up (3 days from now)
                FollowUp.create_from_symptom(record, days_ahead=3)
                
                record_id = str(record.id)
                logger.info(f"Created symptom record {record_id} from chat diagnosis")
                
            except Exception as e:
                logger.error(f"Failed to create symptom record from chat: {e}")
        
        return Response({
            'response': response_text,
            'session_id': str(session_id),
            'source': response_source,  # "rasa" or "llm_fallback"
            'buttons': buttons,  # Interactive buttons from Rasa
            'rasa_available': rasa_service.is_available(),  # Debug info
            'record_id': record_id,  # ID of created symptom record (if any)
            'diagnosis_saved': record_id is not None  # Whether diagnosis was saved to history
        })
    
    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def generate_insights(request):
    """
    Generate top 3 health insights for current session
    POST /api/chat/insights/
    """
    session_id = request.data.get('session_id')
    symptoms = request.data.get('symptoms', [])
    disease = request.data.get('disease', '')
    
    if not session_id:
        return Response(
            {'error': 'session_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        session = ChatSession.objects.get(id=session_id, student=request.user)
        
        # Delete old insights for this session
        HealthInsight.objects.filter(student=request.user, session_id=session_id).delete()
        
        # Get ML predictions for context
        predictor = get_ml_predictor()
        prediction_results = predictor.predict(symptoms)
        
        # Generate new insights using LLM service
        insights_data = ai_generator.generate_health_insights(
            symptoms=symptoms,
            predictions=prediction_results,
            chat_summary=session.metadata.get('topics_discussed', '')
        )
        
        # Save top 3 insights
        insights = []
        for insight_data in insights_data[:3]:
            insight = HealthInsight.objects.create(
                student=request.user,
                session_id=session_id,
                insight_text=insight_data['text'],
                references=[],  # LLM doesn't provide references yet
                reliability_score=insight_data['reliability_score']
            )
            insights.append(insight)
        
        # Update session
        session.insights_generated_count = len(insights)
        session.save()
        
        serializer = HealthInsightSerializer(insights, many=True)
        return Response(serializer.data)
    
    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent])
def end_chat_session(request):
    """
    End chat session and calculate duration
    POST /api/chat/end/
    """
    session_id = request.data.get('session_id')
    
    if not session_id:
        return Response(
            {'error': 'session_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        session = ChatSession.objects.get(id=session_id, student=request.user)
        
        if session.ended_at:
            return Response(
                {'error': 'Session already ended'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        session.ended_at = timezone.now()
        session.duration_seconds = int((session.ended_at - session.started_at).total_seconds())
        session.save()
        
        return Response({
            'message': 'Session ended',
            'duration_seconds': session.duration_seconds
        })
    
    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# Clinic Staff Dashboard Views
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def clinic_dashboard(request):
    """
    Get clinic dashboard overview with statistics
    GET /api/staff/dashboard/
    """
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)
    
    # Overall statistics
    total_students = User.objects.filter(role='student').count()
    
    students_today = SymptomRecord.objects.filter(
        created_at__date=today
    ).values('student').distinct().count()
    
    students_7days = SymptomRecord.objects.filter(
        created_at__date__gte=seven_days_ago
    ).values('student').distinct().count()
    
    students_30days = SymptomRecord.objects.filter(
        created_at__date__gte=thirty_days_ago
    ).values('student').distinct().count()
    
    pending_referrals = SymptomRecord.objects.filter(
        requires_referral=True,
        referral_triggered=False
    ).count()
    
    # Department breakdown - Calculate from real data
    departments = User.objects.filter(role='student').values_list('department', flat=True).distinct()
    dept_breakdown = []
    
    for dept in departments:
        if not dept:  # Skip None/empty departments
            continue
            
        total_in_dept = User.objects.filter(role='student', department=dept).count()
        students_with_symptoms = SymptomRecord.objects.filter(
            student__department=dept,
            created_at__date__gte=thirty_days_ago
        ).values('student').distinct().count()
        
        dept_breakdown.append({
            'department': dept,
            'total_students': total_in_dept,
            'students_with_symptoms': students_with_symptoms,
            'percentage': round((students_with_symptoms / total_in_dept * 100), 1) if total_in_dept > 0 else 0
        })
    
    # Sort by students with symptoms (descending)
    dept_breakdown.sort(key=lambda x: x['students_with_symptoms'], reverse=True)
    
    # Recent symptom records
    recent_symptoms = SymptomRecord.objects.select_related('student').order_by('-created_at')[:10]
    
    # Top insight (most common disease in last 30 days)
    top_disease = SymptomRecord.objects.filter(
        created_at__date__gte=thirty_days_ago
    ).values('predicted_disease').annotate(
        count=Count('id')
    ).order_by('-count').first()
    
    top_insight = f"{top_disease['predicted_disease']} ({top_disease['count']} cases this month)" if top_disease else 'No consultations yet'
    
    # Prepare response
    data = {
        'total_students': total_students,
        'students_with_symptoms_today': students_today,
        'students_with_symptoms_7days': students_7days,
        'students_with_symptoms_30days': students_30days,
        'top_insight': top_insight,
        'department_breakdown': dept_breakdown,
        'recent_symptoms': SymptomRecordSerializer(recent_symptoms, many=True).data,
        'pending_referrals': pending_referrals
    }
    
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def student_directory(request):
    """
    Get filtered list of students with health records
    GET /api/staff/students/
    """
    from django.db.models import Prefetch, Avg
    
    queryset = User.objects.filter(role='student').prefetch_related(
        'symptom_records',
        'medications',
        'follow_ups'
    )
    
    # Filters
    department = request.query_params.get('department')
    search = request.query_params.get('search')
    has_symptoms = request.query_params.get('has_symptoms')
    
    if department:
        queryset = queryset.filter(department__icontains=department)
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) | Q(school_id__icontains=search)
        )
    
    if has_symptoms == 'true':
        queryset = queryset.filter(symptom_records__isnull=False).distinct()
    elif has_symptoms == 'false':
        queryset = queryset.filter(symptom_records__isnull=True)
    
    # Build enriched student data
    students_data = []
    for student in queryset:
        # Get symptom records
        recent_symptoms = student.symptom_records.order_by('-created_at')[:5]
        last_visit = recent_symptoms.first().created_at if recent_symptoms.exists() else None
        
        # Get medications
        active_meds = student.medications.filter(is_active=True)
        
        # Calculate adherence
        med_logs = MedicationLog.objects.filter(medication__student=student)
        total_logs = med_logs.count()
        taken_logs = med_logs.filter(status='taken').count()
        adherence_rate = round((taken_logs / total_logs * 100) if total_logs > 0 else 100, 1)
        
        # Get follow-ups
        pending_followups = student.follow_ups.filter(status='pending').exists()
        
        student_data = {
            'id': student.id,
            'name': student.name,
            'school_id': student.school_id,
            'department': student.department,
            'total_visits': student.symptom_records.count(),
            'last_visit': last_visit.isoformat() if last_visit else None,
            'on_medication': active_meds.exists(),
            'medication_count': active_meds.count(),
            'adherence_rate': adherence_rate,
            'pending_followup': pending_followups,
            'recent_symptoms': recent_symptoms.exists(),
            'recent_symptom_reports': SymptomRecordSerializer(recent_symptoms, many=True).data,
            'medications': MedicationSerializer(active_meds, many=True).data
        }
        
        students_data.append(student_data)
    
    return Response({'students': students_data})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def export_report(request):
    """
    Export symptom data to Excel or CSV format
    GET /api/staff/export/?format=csv|excel&start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
    
    Query Parameters:
        - format: 'csv' or 'excel' (default: csv)
        - start_date: Filter records from this date (optional)
        - end_date: Filter records until this date (optional)
        - department: Filter by department (optional)
        - disease: Filter by predicted disease (optional)
    """
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get query parameters
    export_format = request.query_params.get('format', 'csv').lower()
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    department = request.query_params.get('department')
    disease = request.query_params.get('disease')
    
    # Build queryset
    queryset = SymptomRecord.objects.select_related('student').all()
    
    if start_date:
        queryset = queryset.filter(created_at__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__lte=end_date)
    if department:
        queryset = queryset.filter(student__department=department)
    if disease:
        queryset = queryset.filter(predicted_disease__icontains=disease)
    
    queryset = queryset.order_by('-created_at')
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Symptom Records"
            
            # Define headers
            headers = [
                'Date', 'Time', 'Student ID', 'Student Name', 'Department',
                'Symptoms', 'Duration (days)', 'Severity', 'Predicted Disease',
                'Confidence', 'ICD-10 Code', 'Communicable', 'Acute',
                'Requires Referral'
            ]
            
            # Style header row
            header_fill = PatternFill(start_color='006B3F', end_color='006B3F', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for row_num, record in enumerate(queryset, 2):
                severity_map = {1: 'Mild', 2: 'Moderate', 3: 'Severe'}
                
                ws.cell(row=row_num, column=1).value = record.created_at.strftime('%Y-%m-%d')
                ws.cell(row=row_num, column=2).value = record.created_at.strftime('%H:%M:%S')
                ws.cell(row=row_num, column=3).value = record.student.school_id
                ws.cell(row=row_num, column=4).value = record.student.name
                ws.cell(row=row_num, column=5).value = record.student.department
                ws.cell(row=row_num, column=6).value = ', '.join(record.symptoms)
                ws.cell(row=row_num, column=7).value = record.duration_days
                ws.cell(row=row_num, column=8).value = severity_map.get(record.severity, 'Unknown')
                ws.cell(row=row_num, column=9).value = record.predicted_disease
                ws.cell(row=row_num, column=10).value = f"{record.confidence_score:.1%}" if record.confidence_score else 'N/A'
                ws.cell(row=row_num, column=11).value = record.icd10_code or 'N/A'
                ws.cell(row=row_num, column=12).value = 'Yes' if record.is_communicable else 'No'
                ws.cell(row=row_num, column=13).value = 'Yes' if record.is_acute else 'No'
                ws.cell(row=row_num, column=14).value = 'Yes' if record.requires_referral else 'No'
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                for row in ws[column_letter]:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="cpsu_health_report_{timestamp}.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl library. Install with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    else:  # CSV format
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cpsu_health_report_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Date', 'Time', 'Student ID', 'Student Name', 'Department',
            'Symptoms', 'Duration (days)', 'Severity', 'Predicted Disease',
            'Confidence', 'ICD-10 Code', 'Communicable', 'Acute',
            'Requires Referral'
        ])
        
        # Write data rows
        severity_map = {1: 'Mild', 2: 'Moderate', 3: 'Severe'}
        
        for record in queryset:
            writer.writerow([
                record.created_at.strftime('%Y-%m-%d'),
                record.created_at.strftime('%H:%M:%S'),
                record.student.school_id,
                record.student.name,
                record.student.department,
                ', '.join(record.symptoms),
                record.duration_days,
                severity_map.get(record.severity, 'Unknown'),
                record.predicted_disease,
                f"{record.confidence_score:.1%}" if record.confidence_score else 'N/A',
                record.icd10_code or 'N/A',
                'Yes' if record.is_communicable else 'No',
                'Yes' if record.is_acute else 'No',
                'Yes' if record.requires_referral else 'No'
            ])
        
        return response


# ============================================================================
# Audit Log Views (Staff Only)
# ============================================================================

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    View audit logs (staff only, read-only)
    GET /api/audit/
    """
    queryset = AuditLog.objects.all().select_related('user').order_by('-timestamp')
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsClinicStaff]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by action type
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by user
        school_id = self.request.query_params.get('school_id')
        if school_id:
            queryset = queryset.filter(user__school_id=school_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        
        return queryset


# ============================================================================
# Emergency SOS System
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def trigger_emergency(request):
    """
    Trigger emergency SOS alert
    POST /api/emergency/trigger/
    
    Immediately notifies all clinic staff
    """
    serializer = EmergencyTriggerSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Create emergency alert
    emergency = EmergencyAlert.objects.create(
        student=request.user,
        location=serializer.validated_data['location'],
        symptoms=serializer.validated_data.get('symptoms', []),
        description=serializer.validated_data.get('description', ''),
        status='active',
        priority=100  # All emergencies are critical
    )
    
    # Log the emergency
    AuditLog.objects.create(
        user=request.user,
        action='emergency_triggered',
        details={
            'emergency_id': str(emergency.id),
            'location': emergency.location,
            'symptoms_count': len(emergency.symptoms)
        }
    )
    
    # TODO: Send real-time notifications to staff
    # This will be implemented with WebSockets or push notifications
    # For now, staff can poll /api/emergency/active/
    
    return Response({
        'status': 'emergency_triggered',
        'message': 'Help is on the way! Stay where you are.',
        'emergency_id': emergency.id,
        'emergency': EmergencyAlertSerializer(emergency).data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def emergency_active(request):
    """
    Get active emergencies
    GET /api/emergency/active/
    
    Students: See their own active emergencies
    Staff: See all active emergencies
    """
    if request.user.role == 'staff':
        # Staff can see all active emergencies
        emergencies = EmergencyAlert.objects.filter(
            status__in=['active', 'responding']
        ).select_related('student', 'responded_by')
    else:
        # Students see only their own
        emergencies = EmergencyAlert.objects.filter(
            student=request.user,
            status__in=['active', 'responding']
        )
    
    serializer = EmergencyAlertSerializer(emergencies, many=True)
    return Response({
        'count': emergencies.count(),
        'emergencies': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def emergency_history(request):
    """
    Get emergency history
    GET /api/emergency/history/
    
    Students: Their own history
    Staff: All emergencies
    """
    if request.user.role == 'staff':
        emergencies = EmergencyAlert.objects.all().select_related('student', 'responded_by')
    else:
        emergencies = EmergencyAlert.objects.filter(student=request.user)
    
    # Pagination
    page_size = int(request.GET.get('page_size', 20))
    emergencies = emergencies[:page_size]
    
    serializer = EmergencyAlertSerializer(emergencies, many=True)
    return Response({
        'count': emergencies.count(),
        'emergencies': serializer.data
    })


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def emergency_respond(request, emergency_id):
    """
    Staff responds to emergency
    PATCH /api/emergency/<id>/respond/
    """
    try:
        emergency = EmergencyAlert.objects.get(id=emergency_id)
    except EmergencyAlert.DoesNotExist:
        return Response({'error': 'Emergency not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Update status to responding if not already
    if emergency.status == 'active':
        emergency.status = 'responding'
        emergency.responded_by = request.user
        emergency.response_time = timezone.now()
        emergency.save()
        
        # Log response
        AuditLog.objects.create(
            user=request.user,
            action='emergency_responded',
            details={
                'emergency_id': str(emergency.id),
                'student': emergency.student.school_id,
                'response_time_minutes': emergency.response_time_minutes
            }
        )
    
    serializer = EmergencyAlertSerializer(emergency)
    return Response({
        'message': 'Emergency status updated to responding',
        'emergency': serializer.data
    })


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def emergency_resolve(request, emergency_id):
    """
    Staff resolves emergency
    PATCH /api/emergency/<id>/resolve/
    """
    try:
        emergency = EmergencyAlert.objects.get(id=emergency_id)
    except EmergencyAlert.DoesNotExist:
        return Response({'error': 'Emergency not found'}, status=status.HTTP_404_NOT_FOUND)
    
    resolution_notes = request.data.get('notes', '')
    is_false_alarm = request.data.get('false_alarm', False)
    
    if is_false_alarm:
        emergency.status = 'false_alarm'
    else:
        emergency.status = 'resolved'
    
    emergency.resolved_at = timezone.now()
    emergency.resolution_notes = resolution_notes
    emergency.responded_by = request.user
    
    if not emergency.response_time:
        emergency.response_time = timezone.now()
    
    emergency.save()
    
    # Log resolution
    AuditLog.objects.create(
        user=request.user,
        action='emergency_resolved',
        details={
            'emergency_id': str(emergency.id),
            'student': emergency.student.school_id,
            'resolution': emergency.status,
            'response_time_minutes': emergency.response_time_minutes
        }
    )
    
    serializer = EmergencyAlertSerializer(emergency)
    return Response({
        'message': f'Emergency {emergency.get_status_display()}',
        'emergency': serializer.data
    })


# ============================================================================
# Medication Management System
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_list(request):
    """
    Get medications for current user
    GET /api/medications/
    
    Students: Their own medications
    Staff: Can filter by student_id
    """
    if request.user.role == 'staff':
        # Staff can query specific student
        student_id = request.GET.get('student_id')
        if student_id:
            medications = Medication.objects.filter(
                student__school_id=student_id
            ).select_related('student', 'prescribed_by', 'symptom_record')
        else:
            # All medications (for staff dashboard)
            medications = Medication.objects.all().select_related(
                'student', 'prescribed_by'
            )[:50]  # Limit to recent 50
    else:
        # Students see only their own
        medications = Medication.objects.filter(
            student=request.user
        ).select_related('prescribed_by', 'symptom_record')
    
    # Filter by active status
    active_only = request.GET.get('active_only', 'false').lower() == 'true'
    if active_only:
        medications = medications.filter(is_active=True)
    
    serializer = MedicationSerializer(medications, many=True)
    return Response({
        'count': medications.count(),
        'medications': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsClinicStaff])
def medication_create(request):
    """
    Staff prescribes medication to student
    POST /api/medications/
    """
    serializer = MedicationCreateSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Create medication
    medication = serializer.save(prescribed_by=request.user)
    
    # Auto-generate medication logs based on schedule
    from datetime import datetime, timedelta
    
    current_date = medication.start_date
    while current_date <= medication.end_date:
        for time_str in medication.schedule_times:
            # Parse time (HH:MM format)
            try:
                scheduled_time = datetime.strptime(time_str, '%H:%M').time()
                MedicationLog.objects.create(
                    medication=medication,
                    scheduled_date=current_date,
                    scheduled_time=scheduled_time,
                    status='pending'
                )
            except ValueError:
                pass  # Skip invalid time formats
        
        current_date += timedelta(days=1)
    
    # Log the prescription
    AuditLog.objects.create(
        user=request.user,
        action='medication_prescribed',
        details={
            'medication_id': str(medication.id),
            'student': medication.student.school_id,
            'medication_name': medication.name,
            'duration_days': (medication.end_date - medication.start_date).days + 1
        }
    )
    
    return Response({
        'message': 'Medication prescribed successfully',
        'medication': MedicationSerializer(medication).data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_detail(request, medication_id):
    """
    Get medication details
    GET /api/medications/<id>/
    """
    try:
        medication = Medication.objects.select_related(
            'student', 'prescribed_by', 'symptom_record'
        ).prefetch_related('logs').get(id=medication_id)
    except Medication.DoesNotExist:
        return Response({'error': 'Medication not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Permission check
    if request.user.role != 'staff' and medication.student != request.user:
        return Response(
            {'error': 'You do not have permission to view this medication'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    serializer = MedicationSerializer(medication)
    return Response(serializer.data)


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def medication_update(request, medication_id):
    """
    Update medication (staff only)
    PATCH /api/medications/<id>/
    """
    try:
        medication = Medication.objects.get(id=medication_id)
    except Medication.DoesNotExist:
        return Response({'error': 'Medication not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Update fields
    allowed_fields = ['dosage', 'frequency', 'instructions', 'is_active', 'end_date']
    for field in allowed_fields:
        if field in request.data:
            setattr(medication, field, request.data[field])
    
    medication.save()
    
    return Response({
        'message': 'Medication updated',
        'medication': MedicationSerializer(medication).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_logs_today(request):
    """
    Get today's medication schedule for student
    GET /api/medications/logs/today/
    """
    today = timezone.now().date()
    
    if request.user.role == 'student':
        # Get student's medications
        medications = Medication.objects.filter(
            student=request.user,
            is_active=True
        )
        
        logs = MedicationLog.objects.filter(
            medication__in=medications,
            scheduled_date=today
        ).select_related('medication').order_by('scheduled_time')
        
    else:
        # Staff sees all for the day
        logs = MedicationLog.objects.filter(
            scheduled_date=today
        ).select_related('medication', 'medication__student')[:100]
    
    serializer = MedicationLogSerializer(logs, many=True)
    return Response({
        'date': today,
        'count': logs.count(),
        'logs': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def medication_log_mark_taken(request, log_id):
    """
    Mark medication dose as taken
    POST /api/medications/logs/<id>/taken/
    """
    try:
        log = MedicationLog.objects.select_related('medication').get(id=log_id)
    except MedicationLog.DoesNotExist:
        return Response({'error': 'Log not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Permission check
    if request.user.role != 'staff' and log.medication.student != request.user:
        return Response(
            {'error': 'Permission denied'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    notes = request.data.get('notes', '')
    log.mark_as_taken(notes=notes)
    
    return Response({
        'message': 'Marked as taken',
        'log': MedicationLogSerializer(log).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_adherence(request):
    """
    Get medication adherence statistics
    GET /api/medications/adherence/
    """
    if request.user.role == 'student':
        medications = Medication.objects.filter(student=request.user, is_active=True)
    else:
        student_id = request.GET.get('student_id')
        if student_id:
            medications = Medication.objects.filter(
                student__school_id=student_id,
                is_active=True
            )
        else:
            return Response({'error': 'student_id required for staff'}, status=400)
    
    stats = []
    for med in medications:
        total_logs = med.logs.exclude(status='pending').count()
        if total_logs > 0:
            taken_count = med.logs.filter(status='taken').count()
            missed_count = med.logs.filter(status='missed').count()
            adherence_rate = (taken_count / total_logs) * 100
            
            stats.append({
                'medication_id': str(med.id),
                'medication_name': med.name,
                'total_doses': total_logs,
                'taken': taken_count,
                'missed': missed_count,
                'adherence_rate': round(adherence_rate, 1),
                'days_remaining': med.days_remaining
            })
    
    # Overall adherence
    total_all = sum(s['total_doses'] for s in stats)
    taken_all = sum(s['taken'] for s in stats)
    overall_rate = (taken_all / total_all * 100) if total_all > 0 else 0
    
    return Response({
        'overall_adherence_rate': round(overall_rate, 1),
        'medications': stats
    })


# ============================================================================
# Follow-Up Views
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def followup_list(request):
    """
    Get follow-ups for current user (students) or all/specific student (staff)
    GET /api/followups/
    Query params: status, student_id (staff only)
    """
    if request.user.role == 'student':
        followups = FollowUp.objects.filter(student=request.user)
    else:
        # Staff can see all or filter by student
        student_id = request.GET.get('student_id')
        if student_id:
            followups = FollowUp.objects.filter(student__school_id=student_id)
        else:
            followups = FollowUp.objects.all()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        followups = followups.filter(status=status_filter)
    
    # Auto-check and update overdue status
    for followup in followups:
        followup.check_overdue()
    
    serializer = FollowUpSerializer(followups, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def followup_pending(request):
    """
    Get pending follow-ups (including overdue) for current user
    GET /api/followups/pending/
    """
    followups = FollowUp.objects.filter(
        student=request.user,
        status__in=['pending', 'overdue']
    ).order_by('scheduled_date')
    
    # Check for overdue
    for followup in followups:
        followup.check_overdue()
    
    serializer = FollowUpSerializer(followups, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent])
def followup_respond(request, pk):
    """
    Submit response to follow-up
    POST /api/followups/<id>/respond/
    """
    try:
        followup = FollowUp.objects.get(pk=pk, student=request.user)
    except FollowUp.DoesNotExist:
        return Response({'error': 'Follow-up not found'}, status=404)
    
    if followup.status == 'completed':
        return Response({'error': 'Follow-up already completed'}, status=400)
    
    serializer = FollowUpResponseSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)
    
    # Update follow-up with response
    followup.outcome = serializer.validated_data['outcome']
    followup.notes = serializer.validated_data.get('notes', '')
    followup.still_experiencing_symptoms = serializer.validated_data['still_experiencing_symptoms']
    followup.new_symptoms = serializer.validated_data.get('new_symptoms', [])
    followup.response_date = timezone.now()
    followup.status = 'completed'
    followup.save()
    
    # Auto-flag for appointment if condition worsened
    if followup.outcome == 'worse':
        followup.requires_appointment = True
        followup.save()
    
    return Response({
        'message': 'Follow-up response submitted',
        'followup': FollowUpSerializer(followup).data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def followup_review(request, pk):
    """
    Staff review of follow-up response
    POST /api/followups/<id>/review/
    """
    try:
        followup = FollowUp.objects.get(pk=pk)
    except FollowUp.DoesNotExist:
        return Response({'error': 'Follow-up not found'}, status=404)
    
    review_notes = request.data.get('review_notes', '')
    requires_appointment = request.data.get('requires_appointment', False)
    
    followup.reviewed_by = request.user
    followup.review_notes = review_notes
    followup.requires_appointment = requires_appointment
    followup.save()
    
    return Response({
        'message': 'Follow-up reviewed',
        'followup': FollowUpSerializer(followup).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def followup_needs_review(request):
    """
    Get follow-ups that need staff review
    GET /api/followups/needs-review/
    """
    # Get all follow-ups that need attention
    followups = FollowUp.objects.select_related('student', 'reviewed_by').order_by('-scheduled_date')
    
    # Enrich with student data
    data = []
    for followup in followups:
        followup_data = FollowUpSerializer(followup).data
        followup_data['student_name'] = followup.student.name
        followup_data['student_school_id'] = followup.student.school_id
        followup_data['student_department'] = followup.student.department
        followup_data['reviewed_by_name'] = followup.reviewed_by.name if followup.reviewed_by else None
        data.append(followup_data)
    
    return Response(data)


# ============================================================================
# Analytics Views (Real Data)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def staff_analytics(request):
    """
    Get comprehensive analytics data for charts
    GET /api/staff/analytics/?period=7d
    
    Periods: 7d, 30d, 90d, 1y
    """
    period = request.query_params.get('period', '30d')
    
    # Calculate date range
    today = timezone.now().date()
    if period == '7d':
        start_date = today - timedelta(days=7)
    elif period == '30d':
        start_date = today - timedelta(days=30)
    elif period == '90d':
        start_date = today - timedelta(days=90)
    elif period == '1y':
        start_date = today - timedelta(days=365)
    else:
        start_date = today - timedelta(days=30)
    
    # Summary stats
    total_consultations = SymptomRecord.objects.filter(created_at__date__gte=start_date).count()
    unique_patients = SymptomRecord.objects.filter(created_at__date__gte=start_date).values('student').distinct().count()
    emergency_alerts = EmergencyAlert.objects.filter(created_at__date__gte=start_date).count()
    prescriptions = Medication.objects.filter(created_at__date__gte=start_date).count()
    
    # Top 10 diagnosed conditions
    top_conditions = SymptomRecord.objects.filter(
        created_at__date__gte=start_date
    ).values('predicted_disease').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Consultation trends (daily counts)
    consultation_trends = []
    current_date = start_date
    while current_date <= today:
        count = SymptomRecord.objects.filter(created_at__date=current_date).count()
        consultation_trends.append({
            'date': current_date.isoformat(),
            'count': count
        })
        current_date += timedelta(days=1)
    
    # Consultations by department
    dept_breakdown = SymptomRecord.objects.filter(
        created_at__date__gte=start_date
    ).values('student__department').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Symptom severity distribution (using actual severity field: 1=Mild, 2=Moderate, 3=Severe)
    severity_distribution = {
        'mild': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=1).count(),
        'moderate': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=2).count(),
        'severe': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=3).count(),
    }
    
    # Most common symptoms
    from collections import Counter
    
    symptom_records = SymptomRecord.objects.filter(created_at__date__gte=start_date)
    all_symptoms = []
    for record in symptom_records:
        if record.symptoms:  # Changed from symptoms_reported to symptoms
            all_symptoms.extend(record.symptoms)
    
    symptom_counter = Counter(all_symptoms)
    common_symptoms = [
        {
            'symptom': symptom,
            'count': count,
            'percentage': round((count / len(all_symptoms) * 100) if all_symptoms else 0, 1)
        }
        for symptom, count in symptom_counter.most_common(10)
    ]
    
    data = {
        'period': period,
        'summary': {
            'total_consultations': total_consultations,
            'unique_patients': unique_patients,
            'emergency_alerts': emergency_alerts,
            'prescriptions': prescriptions
        },
        'top_conditions': list(top_conditions),
        'consultation_trends': consultation_trends,
        'department_breakdown': list(dept_breakdown),
        'severity_distribution': severity_distribution,
        'common_symptoms': common_symptoms
    }
    
    return Response(data)
